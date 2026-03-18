"""Generate pipeline overview Excel file."""
import sqlite3, sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

conn = sqlite3.connect('data/processed/pipeline.db', timeout=10)
conn.row_factory = sqlite3.Row
wb = openpyxl.Workbook()

thin = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
hdr_font = Font(bold=True, color='FFFFFF', size=11)
num_fmt = '#,##0'

def styled_header(ws, headers, fill_color):
    fill = PatternFill('solid', fgColor=fill_color)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = hdr_font
        cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

# ─── Лист 1: Сводка ───
ws = wb.active
ws.title = 'Сводка по таблицам'
styled_header(ws, ['Таблица', 'Всего', 'legacy', 'osm', 'restoclub', 'afisha'], '4472C4')

summary = [
    ('restaurants', 'SELECT COUNT(*) FROM restaurants',
     'SELECT source, COUNT(*) FROM restaurants GROUP BY source'),
    ('photos', 'SELECT COUNT(*) FROM photos',
     'SELECT source, COUNT(*) FROM photos GROUP BY source'),
    ('restaurant_cuisines', 'SELECT COUNT(*) FROM restaurant_cuisines',
     'SELECT r.source, COUNT(*) FROM restaurant_cuisines rc JOIN restaurants r ON rc.restaurant_id=r.id GROUP BY r.source'),
    ('working_hours (рест.)', 'SELECT COUNT(DISTINCT restaurant_id) FROM working_hours',
     'SELECT r.source, COUNT(DISTINCT w.restaurant_id) FROM working_hours w JOIN restaurants r ON w.restaurant_id=r.id GROUP BY r.source'),
    ('dishes', 'SELECT COUNT(*) FROM dishes', None),
    ('cuisines', 'SELECT COUNT(*) FROM cuisines', None),
    ('cities', 'SELECT COUNT(*) FROM cities', None),
    ('districts', 'SELECT COUNT(*) FROM districts', None),
]

for i, (name, total_q, src_q) in enumerate(summary, 2):
    total = conn.execute(total_q).fetchone()[0]
    ws.cell(i, 1, name).border = thin
    c = ws.cell(i, 2, total); c.border = thin; c.number_format = num_fmt

    if src_q:
        by_src = {r[0]: r[1] for r in conn.execute(src_q)}
        for j, src in enumerate(['legacy', 'osm', 'restoclub', 'afisha'], 3):
            v = by_src.get(src, '')
            c = ws.cell(i, j, v if v else '')
            c.border = thin
            if isinstance(v, int):
                c.number_format = num_fmt
    else:
        for j in range(3, 7):
            ws.cell(i, j, '').border = thin

for w, col in [(25,1),(12,2),(12,3),(12,4),(12,5),(12,6)]:
    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

# ─── Лист 2: Города ───
ws2 = wb.create_sheet('Города')
city_headers = ['#', 'Город', 'Всего рест.', 'legacy', 'osm', 'restoclub', 'afisha', 'Фото', 'С расписанием']
styled_header(ws2, city_headers, '548235')

city_rows = conn.execute("""
    SELECT city,
        COUNT(*) as total,
        SUM(CASE WHEN source='legacy' THEN 1 ELSE 0 END),
        SUM(CASE WHEN source='osm' THEN 1 ELSE 0 END),
        SUM(CASE WHEN source='restoclub' THEN 1 ELSE 0 END),
        SUM(CASE WHEN source='afisha' THEN 1 ELSE 0 END)
    FROM restaurants WHERE city IS NOT NULL AND is_duplicate=0
    GROUP BY city ORDER BY COUNT(*) DESC
""").fetchall()

photo_by_city = {r[0]: r[1] for r in conn.execute(
    "SELECT r.city, COUNT(*) FROM photos p JOIN restaurants r ON p.restaurant_id=r.id WHERE r.city IS NOT NULL GROUP BY r.city"
)}
hours_by_city = {r[0]: r[1] for r in conn.execute(
    "SELECT r.city, COUNT(DISTINCT w.restaurant_id) FROM working_hours w JOIN restaurants r ON w.restaurant_id=r.id WHERE r.city IS NOT NULL GROUP BY r.city"
)}

for i, row in enumerate(city_rows, 2):
    city_name = row[0]
    vals = [i-1, city_name, row[1], row[2], row[3], row[4], row[5],
            photo_by_city.get(city_name, 0), hours_by_city.get(city_name, 0)]
    for c, v in enumerate(vals, 1):
        cell = ws2.cell(i, c, v)
        cell.border = thin
        if isinstance(v, int):
            cell.number_format = num_fmt

for w, col in [(4,1),(22,2),(12,3),(10,4),(10,5),(10,6),(10,7),(10,8),(14,9)]:
    ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:I{len(city_rows)+1}'

# ─── Листы 3-4: Restoclub/Afisha по 100 ───
rest_cols = ['#','Название','Город','Адрес','Метро','Широта','Долгота','Телефон',
             'Сайт','Кухни','Ценовой диапазон','Средний чек','Рейтинг','Отзывов',
             'Часы работы','Фичи','Фото','Описание']
rest_widths = [4,28,16,32,16,10,10,18,30,28,16,11,8,8,35,28,6,45]

for source, color in [('restoclub', 'ED7D31'), ('afisha', '4472C4')]:
    ws3 = wb.create_sheet(f'{source.capitalize()} (100)')
    styled_header(ws3, rest_cols, color)

    rows = conn.execute("""
        SELECT r.name, r.city, r.address, r.metro_station, r.lat, r.lng,
               r.phone, r.website, r.cuisine, r.price_range, r.average_bill,
               r.rating, r.review_count, r.opening_hours, r.features, r.description,
               (SELECT COUNT(*) FROM photos p WHERE p.restaurant_id = r.id) as photo_count
        FROM restaurants r WHERE r.source = ? ORDER BY RANDOM() LIMIT 100
    """, (source,)).fetchall()

    for i, r in enumerate(rows, 2):
        desc = (r['description'] or '')[:500]
        vals = [i-1, r['name'], r['city'], r['address'], r['metro_station'],
                r['lat'], r['lng'], r['phone'], r['website'], r['cuisine'],
                r['price_range'], r['average_bill'], r['rating'], r['review_count'],
                r['opening_hours'], r['features'], r['photo_count'], desc]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(i, c, v)
            cell.border = thin
            cell.alignment = Alignment(vertical='top', wrap_text=(c in [10,15,16,18]))

    for c, w in enumerate(rest_widths, 1):
        ws3.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
    ws3.freeze_panes = 'A2'
    ws3.auto_filter.ref = f'A1:R101'

# ─── Лист 5: Заполненность полей ───
ws5 = wb.create_sheet('Заполненность полей')
fh = ['Поле', 'legacy', '%', 'osm', '%', 'restoclub', '%', 'afisha', '%']
styled_header(ws5, fh, '7030A0')

fields = [
    ('Адрес', "address IS NOT NULL AND address != ''"),
    ('Координаты', 'lat IS NOT NULL'),
    ('Телефон', "phone IS NOT NULL AND phone != ''"),
    ('Описание', "description IS NOT NULL AND description != ''"),
    ('Кухни', "cuisine IS NOT NULL AND cuisine != '[]'"),
    ('Рейтинг', 'rating IS NOT NULL AND rating > 0'),
    ('Часы работы', "opening_hours IS NOT NULL AND opening_hours != ''"),
    ('Фичи', "features IS NOT NULL AND features != '[]'"),
    ('Метро', 'metro_station IS NOT NULL'),
    ('Средний чек', 'average_bill IS NOT NULL'),
]

for i, (label, cond) in enumerate(fields, 2):
    ws5.cell(i, 1, label).border = thin
    col = 2
    for src in ['legacy', 'osm', 'restoclub', 'afisha']:
        total = conn.execute('SELECT COUNT(*) FROM restaurants WHERE source=?', (src,)).fetchone()[0]
        cnt = conn.execute(f'SELECT COUNT(*) FROM restaurants WHERE source=? AND {cond}', (src,)).fetchone()[0]
        pct = cnt / total * 100 if total else 0
        c1 = ws5.cell(i, col, cnt); c1.border = thin; c1.number_format = num_fmt
        c2 = ws5.cell(i, col + 1, f'{pct:.1f}%'); c2.border = thin
        col += 2

for c, w in [(1,16),(2,10),(3,8),(4,10),(5,8),(6,10),(7,9),(8,10),(9,8)]:
    ws5.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w
ws5.freeze_panes = 'A2'

out = 'data/processed/pipeline_overview.xlsx'
wb.save(out)
print(f'Saved: {os.path.abspath(out)}')
conn.close()
