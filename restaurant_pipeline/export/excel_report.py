"""
Полная сводка pipeline данных в Excel.

Генерирует файл data/processed/pipeline_report.xlsx с листами:
  1. Сводка — общая статистика
  2. Рестораны — топ-данные по ресторанам
  3. Блюда — все блюда с КБЖУ и аллергенами
  4. Города — статистика по городам
  5. Районы — привязка к районам
  6. Кухни — распределение по кухням
  7. Фичи — какие фичи у ресторанов
  8. Рабочие часы — пример структурированных часов
  9. Фото — статистика фотографий
  10. Отзывы — выборка отзывов
"""
import sys
import os
import json
from datetime import datetime

if sys.platform == 'win32':
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..'))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from utils.db import get_connection


# Styles
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FILL_GREEN = PatternFill(start_color='548235', end_color='548235', fill_type='solid')
HEADER_FILL_ORANGE = PatternFill(start_color='ED7D31', end_color='ED7D31', fill_type='solid')
STAT_LABEL_FONT = Font(bold=True, size=11)
STAT_VALUE_FONT = Font(size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(bold=True, size=12, color='2E75B6')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)


def _style_header(ws, row, ncols, fill=None):
    """Apply header styling to a row."""
    f = fill or HEADER_FILL
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = f
        cell.alignment = Alignment(horizontal='center', wrap_text=True)


def _auto_width(ws, min_width=10, max_width=50):
    """Auto-fit column widths."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells[:100]:  # sample first 100 rows
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def _add_borders(ws, start_row, end_row, ncols):
    """Add thin borders to data area."""
    for row in range(start_row, end_row + 1):
        for col in range(1, ncols + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


def generate_excel_report():
    """Generate comprehensive Excel report."""
    print("\n" + "=" * 60)
    print("ГЕНЕРАЦИЯ EXCEL ОТЧЁТА")
    print("=" * 60)

    conn = get_connection()
    wb = Workbook()

    # ========== 1. СВОДКА ==========
    ws = wb.active
    ws.title = "Сводка"
    ws.sheet_properties.tabColor = '4472C4'

    ws.cell(row=1, column=1, value="Menu-Rest Pipeline — Полная сводка").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = STAT_VALUE_FONT
    ws.merge_cells('A1:D1')
    ws.merge_cells('A2:D2')

    stats = []

    # Core counts
    total_rest = conn.execute("SELECT COUNT(*) FROM restaurants WHERE is_duplicate = 0").fetchone()[0]
    total_all = conn.execute("SELECT COUNT(*) FROM restaurants").fetchone()[0]
    duplicates = total_all - total_rest
    cities = conn.execute("SELECT COUNT(*) FROM cities").fetchone()[0]
    cuisines = conn.execute("SELECT COUNT(DISTINCT cuisine) FROM restaurants WHERE cuisine IS NOT NULL AND cuisine != ''").fetchone()[0]
    chains_count = conn.execute("SELECT COUNT(*) FROM restaurant_chains").fetchone()[0]

    stats.append(("ОСНОВНЫЕ ДАННЫЕ", ""))
    stats.append(("Всего записей (с дубликатами)", f"{total_all:,}"))
    stats.append(("Уникальных ресторанов", f"{total_rest:,}"))
    stats.append(("Дубликатов удалено", f"{duplicates:,}"))
    stats.append(("Городов", f"{cities:,}"))
    stats.append(("Кухонь", f"{cuisines:,}"))
    stats.append(("Сетей", f"{chains_count:,}"))
    stats.append(("", ""))

    # Sources
    legacy = conn.execute("SELECT COUNT(*) FROM restaurants WHERE source = 'legacy' AND is_duplicate = 0").fetchone()[0]
    osm = conn.execute("SELECT COUNT(*) FROM restaurants WHERE source = 'osm' AND is_duplicate = 0").fetchone()[0]
    stats.append(("ИСТОЧНИКИ", ""))
    stats.append(("Legacy (MySQL дамп)", f"{legacy:,}"))
    stats.append(("OSM (Overpass API)", f"{osm:,}"))
    stats.append(("", ""))

    # Enrichment
    with_desc = conn.execute("SELECT COUNT(*) FROM restaurants WHERE description IS NOT NULL AND description != '' AND is_duplicate = 0").fetchone()[0]
    with_hours = conn.execute("SELECT COUNT(DISTINCT restaurant_id) FROM working_hours").fetchone()[0]
    with_district = conn.execute("SELECT COUNT(*) FROM restaurants WHERE district_id IS NOT NULL AND is_duplicate = 0").fetchone()[0]

    # Features
    try:
        with_features = conn.execute("""
            SELECT COUNT(DISTINCT id) FROM restaurants
            WHERE is_duplicate = 0 AND features IS NOT NULL AND features != '' AND features != '[]'
        """).fetchone()[0]
    except Exception:
        with_features = 0

    # Photos
    photos = conn.execute("SELECT COUNT(*) FROM photos").fetchone()[0]
    rest_with_photos = conn.execute("SELECT COUNT(DISTINCT restaurant_id) FROM photos").fetchone()[0]

    # Dishes
    total_dishes = conn.execute("SELECT COUNT(*) FROM dishes").fetchone()[0]
    with_kbzhu = conn.execute("SELECT COUNT(*) FROM dishes WHERE calories IS NOT NULL").fetchone()[0]
    try:
        with_healthy = conn.execute("SELECT COUNT(*) FROM dishes WHERE is_healthy_choice = 1").fetchone()[0]
    except Exception:
        with_healthy = 0
    try:
        with_allergens = conn.execute("SELECT COUNT(DISTINCT dish_id) FROM dish_allergens").fetchone()[0]
    except Exception:
        with_allergens = 0

    # Reviews
    try:
        reviews = conn.execute("SELECT COUNT(*) FROM reviews").fetchone()[0]
    except Exception:
        reviews = 0

    stats.append(("ОБОГАЩЕНИЕ", ""))
    stats.append(("С описанием", f"{with_desc:,} ({with_desc/total_rest*100:.1f}%)"))
    stats.append(("С рабочими часами", f"{with_hours:,} ({with_hours/total_rest*100:.1f}%)"))
    stats.append(("С районом", f"{with_district:,} ({with_district/total_rest*100:.1f}%)"))
    stats.append(("С фичами (wifi и т.д.)", f"{with_features:,} ({with_features/total_rest*100:.1f}%)"))
    stats.append(("С фотографиями", f"{rest_with_photos:,} ({rest_with_photos/total_rest*100:.1f}%)"))
    stats.append(("", ""))

    stats.append(("ФОТО", ""))
    stats.append(("Всего фотографий", f"{photos:,}"))
    stats.append(("", ""))

    stats.append(("БЛЮДА", ""))
    stats.append(("Всего блюд", f"{total_dishes:,}"))
    stats.append(("С КБЖУ", f"{with_kbzhu:,} ({with_kbzhu/total_dishes*100:.1f}%)" if total_dishes else "0"))
    stats.append(("Healthy choice", f"{with_healthy:,} ({with_healthy/total_dishes*100:.1f}%)" if total_dishes else "0"))
    stats.append(("С аллергенами", f"{with_allergens:,} ({with_allergens/total_dishes*100:.1f}%)" if total_dishes else "0"))
    stats.append(("", ""))

    stats.append(("ОТЗЫВЫ", ""))
    stats.append(("Всего отзывов", f"{reviews:,}"))

    row = 4
    for label, value in stats:
        if value == "" and label:
            ws.cell(row=row, column=1, value=label).font = SUBTITLE_FONT
        else:
            ws.cell(row=row, column=1, value=label).font = STAT_LABEL_FONT
            ws.cell(row=row, column=2, value=value).font = STAT_VALUE_FONT
        row += 1

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25

    # ========== 2. РЕСТОРАНЫ ==========
    print("  Лист: Рестораны...")
    ws2 = wb.create_sheet("Рестораны")
    ws2.sheet_properties.tabColor = '548235'

    headers = ['ID', 'Название', 'Город', 'Адрес', 'Район', 'Кухня', 'Тип',
               'Рейтинг', 'Отзывов', 'Ср. чек', 'Источник', 'Описание (начало)',
               'Координаты', 'Телефон', 'Сайт']
    for col, h in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=h)
    _style_header(ws2, 1, len(headers), HEADER_FILL_GREEN)

    rows = conn.execute("""
        SELECT r.id, r.name,
               COALESCE(c.name, '') as city,
               r.address,
               COALESCE(d.name, '') as district,
               r.cuisine, r.tags,
               r.rating, r.review_count, r.average_bill,
               r.source,
               SUBSTR(r.description, 1, 100) as desc_start,
               CASE WHEN r.lat IS NOT NULL THEN r.lat || ', ' || r.lng ELSE '' END as coords,
               r.phone, r.website
        FROM restaurants r
        LEFT JOIN cities c ON r.city_id = c.id
        LEFT JOIN districts d ON r.district_id = d.id
        WHERE r.is_duplicate = 0
        ORDER BY r.rating DESC NULLS LAST, r.review_count DESC NULLS LAST
        LIMIT 10000
    """).fetchall()

    for i, row_data in enumerate(rows, 2):
        for col, val in enumerate(row_data, 1):
            ws2.cell(row=i, column=col, value=val)

    _auto_width(ws2)
    _add_borders(ws2, 1, min(len(rows) + 1, 10001), len(headers))
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"

    # ========== 3. БЛЮДА ==========
    print("  Лист: Блюда...")
    ws3 = wb.create_sheet("Блюда")
    ws3.sheet_properties.tabColor = 'ED7D31'

    headers = ['ID', 'Название', 'Категория', 'Состав', 'Ккал', 'Белки', 'Жиры',
               'Углеводы', 'Вес (г)', 'Healthy', 'Аллергены', 'Ресторан', 'Цена']
    for col, h in enumerate(headers, 1):
        ws3.cell(row=1, column=col, value=h)
    _style_header(ws3, 1, len(headers), HEADER_FILL_ORANGE)

    # Get allergens per dish
    allergen_map = {}
    try:
        allergen_rows = conn.execute("""
            SELECT da.dish_id, da.allergen_slug
            FROM dish_allergens da
        """).fetchall()
        for ar in allergen_rows:
            did = ar[0]
            aname = ar[1]
            allergen_map.setdefault(did, []).append(aname)
    except Exception:
        pass

    dish_rows = conn.execute("""
        SELECT d.id, d.name, d.category,
               SUBSTR(d.composition, 1, 150),
               d.calories, d.protein, d.fat, d.carbs, d.weight,
               COALESCE(d.is_healthy_choice, 0),
               r.name as rest_name, d.price
        FROM dishes d
        LEFT JOIN restaurants r ON d.restaurant_id = r.id
        ORDER BY d.id
    """).fetchall()

    for i, row_data in enumerate(dish_rows, 2):
        vals = list(row_data)
        dish_id = vals[0]
        allergens = ', '.join(allergen_map.get(dish_id, []))
        # Insert allergens before restaurant name
        vals[9] = 'Да' if vals[9] else ''
        out = vals[:10] + [allergens] + vals[10:]
        for col, val in enumerate(out, 1):
            ws3.cell(row=i, column=col, value=val)

    _auto_width(ws3)
    _add_borders(ws3, 1, min(len(dish_rows) + 1, 5001), len(headers))
    ws3.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(dish_rows)+1}"

    # ========== 4. ГОРОДА ==========
    print("  Лист: Города...")
    ws4 = wb.create_sheet("Города")
    ws4.sheet_properties.tabColor = '7030A0'

    headers = ['Город', 'Ресторанов', 'С описанием', 'С рабочими часами',
               'С фото', 'С районом', 'Legacy', 'OSM']
    for col, h in enumerate(headers, 1):
        ws4.cell(row=1, column=col, value=h)
    _style_header(ws4, 1, len(headers))

    city_rows = conn.execute("""
        SELECT c.name,
               COUNT(r.id) as total,
               SUM(CASE WHEN r.description IS NOT NULL AND r.description != '' THEN 1 ELSE 0 END),
               0,
               (SELECT COUNT(DISTINCT p.restaurant_id) FROM photos p
                JOIN restaurants r2 ON p.restaurant_id = r2.id WHERE r2.city_id = c.id AND r2.is_duplicate = 0),
               SUM(CASE WHEN r.district_id IS NOT NULL THEN 1 ELSE 0 END),
               SUM(CASE WHEN r.source = 'legacy' THEN 1 ELSE 0 END),
               SUM(CASE WHEN r.source = 'osm' THEN 1 ELSE 0 END)
        FROM cities c
        JOIN restaurants r ON r.city_id = c.id AND r.is_duplicate = 0
        GROUP BY c.id, c.name
        ORDER BY total DESC
    """).fetchall()

    for i, row_data in enumerate(city_rows, 2):
        for col, val in enumerate(row_data, 1):
            ws4.cell(row=i, column=col, value=val)

    _auto_width(ws4)
    _add_borders(ws4, 1, len(city_rows) + 1, len(headers))
    ws4.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(city_rows)+1}"

    # ========== 5. РАЙОНЫ ==========
    print("  Лист: Районы...")
    ws5 = wb.create_sheet("Районы")
    ws5.sheet_properties.tabColor = 'BF8F00'

    headers = ['Район', 'Город', 'Slug', 'Ресторанов']
    for col, h in enumerate(headers, 1):
        ws5.cell(row=1, column=col, value=h)
    _style_header(ws5, 1, len(headers))

    dist_rows = conn.execute("""
        SELECT d.name, c.name, d.slug, COUNT(r.id)
        FROM districts d
        JOIN cities c ON d.city_id = c.id
        LEFT JOIN restaurants r ON r.district_id = d.id AND r.is_duplicate = 0
        GROUP BY d.id
        ORDER BY COUNT(r.id) DESC
    """).fetchall()

    for i, row_data in enumerate(dist_rows, 2):
        for col, val in enumerate(row_data, 1):
            ws5.cell(row=i, column=col, value=val)

    _auto_width(ws5)
    _add_borders(ws5, 1, len(dist_rows) + 1, len(headers))

    # ========== 6. КУХНИ ==========
    print("  Лист: Кухни...")
    ws6 = wb.create_sheet("Кухни")
    ws6.sheet_properties.tabColor = 'C00000'

    headers = ['Кухня', 'Ресторанов', '% от всех']
    for col, h in enumerate(headers, 1):
        ws6.cell(row=1, column=col, value=h)
    _style_header(ws6, 1, len(headers))

    cuisine_rows = conn.execute("""
        SELECT cuisine, COUNT(*) as cnt
        FROM restaurants
        WHERE is_duplicate = 0 AND cuisine IS NOT NULL AND cuisine != ''
        GROUP BY cuisine
        ORDER BY cnt DESC
    """).fetchall()

    for i, row_data in enumerate(cuisine_rows, 2):
        ws6.cell(row=i, column=1, value=row_data[0])
        ws6.cell(row=i, column=2, value=row_data[1])
        ws6.cell(row=i, column=3, value=f"{row_data[1]/total_rest*100:.1f}%")

    _auto_width(ws6)
    _add_borders(ws6, 1, len(cuisine_rows) + 1, len(headers))

    # ========== 7. ФИЧИ ==========
    print("  Лист: Фичи...")
    ws7 = wb.create_sheet("Фичи")
    ws7.sheet_properties.tabColor = '00B050'

    headers = ['Фича', 'Ресторанов']
    for col, h in enumerate(headers, 1):
        ws7.cell(row=1, column=col, value=h)
    _style_header(ws7, 1, len(headers))

    # Parse JSON features
    feature_counts = {}
    feat_rows = conn.execute("""
        SELECT features FROM restaurants
        WHERE is_duplicate = 0 AND features IS NOT NULL AND features != '' AND features != '[]'
    """).fetchall()

    for fr in feat_rows:
        try:
            feats = json.loads(fr[0]) if isinstance(fr[0], str) else fr[0]
            if isinstance(feats, list):
                for f in feats:
                    fname = f if isinstance(f, str) else str(f)
                    feature_counts[fname] = feature_counts.get(fname, 0) + 1
        except Exception:
            pass

    sorted_features = sorted(feature_counts.items(), key=lambda x: -x[1])
    for i, (fname, cnt) in enumerate(sorted_features, 2):
        ws7.cell(row=i, column=1, value=fname)
        ws7.cell(row=i, column=2, value=cnt)

    _auto_width(ws7)
    _add_borders(ws7, 1, len(sorted_features) + 1, len(headers))

    # ========== 8. РАБОЧИЕ ЧАСЫ ==========
    print("  Лист: Рабочие часы...")
    ws8 = wb.create_sheet("Рабочие часы")
    ws8.sheet_properties.tabColor = '0070C0'

    DAY_NAMES = ['Пн', 'Вт', 'Ср', 'Чт', 'Пт', 'Сб', 'Вс']
    headers = ['Ресторан', 'Город'] + DAY_NAMES
    for col, h in enumerate(headers, 1):
        ws8.cell(row=1, column=col, value=h)
    _style_header(ws8, 1, len(headers))

    # Sample restaurants with working hours
    wh_restaurants = conn.execute("""
        SELECT DISTINCT wh.restaurant_id, r.name, COALESCE(c.name, '')
        FROM working_hours wh
        JOIN restaurants r ON wh.restaurant_id = r.id
        LEFT JOIN cities c ON r.city_id = c.id
        WHERE r.is_duplicate = 0
        ORDER BY r.rating DESC NULLS LAST
        LIMIT 2000
    """).fetchall()

    row_idx = 2
    for rest_id, rest_name, city_name in wh_restaurants:
        ws8.cell(row=row_idx, column=1, value=rest_name)
        ws8.cell(row=row_idx, column=2, value=city_name)

        hours = conn.execute("""
            SELECT day_of_week, open_time, close_time, is_closed
            FROM working_hours WHERE restaurant_id = ?
            ORDER BY day_of_week
        """, (rest_id,)).fetchall()

        for h in hours:
            day = h[0]
            if 0 <= day <= 6:
                col = day + 3
                if h[3]:  # is_closed
                    ws8.cell(row=row_idx, column=col, value='Закрыто')
                elif h[1] and h[2]:
                    ws8.cell(row=row_idx, column=col, value=f"{h[1]}-{h[2]}")

        row_idx += 1

    _auto_width(ws8)
    _add_borders(ws8, 1, row_idx - 1, len(headers))
    ws8.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx-1}"

    # ========== 9. ФОТО ==========
    print("  Лист: Фото...")
    ws9 = wb.create_sheet("Фото")
    ws9.sheet_properties.tabColor = 'FF6600'

    headers = ['Ресторан', 'Город', 'Кол-во фото', 'URL (первое)', 'Источник']
    for col, h in enumerate(headers, 1):
        ws9.cell(row=1, column=col, value=h)
    _style_header(ws9, 1, len(headers))

    photo_rows = conn.execute("""
        SELECT r.name, COALESCE(c.name, ''),
               COUNT(p.id), MIN(p.url), p.source
        FROM photos p
        JOIN restaurants r ON p.restaurant_id = r.id
        LEFT JOIN cities c ON r.city_id = c.id
        WHERE r.is_duplicate = 0
        GROUP BY p.restaurant_id, r.name, c.name, p.source
        ORDER BY COUNT(p.id) DESC
        LIMIT 5000
    """).fetchall()

    for i, row_data in enumerate(photo_rows, 2):
        for col, val in enumerate(row_data, 1):
            ws9.cell(row=i, column=col, value=val)

    _auto_width(ws9)
    _add_borders(ws9, 1, len(photo_rows) + 1, len(headers))
    ws9.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(photo_rows)+1}"

    # ========== 10. ОТЗЫВЫ ==========
    print("  Лист: Отзывы...")
    ws10 = wb.create_sheet("Отзывы")
    ws10.sheet_properties.tabColor = 'A5A5A5'

    try:
        headers = ['Ресторан', 'Автор', 'Рейтинг', 'Текст (начало)', 'Дата']
        for col, h in enumerate(headers, 1):
            ws10.cell(row=1, column=col, value=h)
        _style_header(ws10, 1, len(headers))

        review_rows = conn.execute("""
            SELECT r.name, rv.author_name, rv.rating,
                   SUBSTR(rv.text, 1, 200), rv.date
            FROM reviews rv
            JOIN restaurants r ON rv.restaurant_id = r.id
            WHERE r.is_duplicate = 0
            ORDER BY rv.date DESC NULLS LAST
            LIMIT 5000
        """).fetchall()

        for i, row_data in enumerate(review_rows, 2):
            for col, val in enumerate(row_data, 1):
                ws10.cell(row=i, column=col, value=val)

        _auto_width(ws10)
        _add_borders(ws10, 1, len(review_rows) + 1, len(headers))
    except Exception as e:
        ws10.cell(row=2, column=1, value=f"Нет данных: {e}")

    # ========== 11. АЛЛЕРГЕНЫ ==========
    print("  Лист: Аллергены...")
    ws11 = wb.create_sheet("Аллергены")
    ws11.sheet_properties.tabColor = 'FF0000'

    try:
        headers = ['Аллерген', 'Блюд с этим аллергеном']
        for col, h in enumerate(headers, 1):
            ws11.cell(row=1, column=col, value=h)
        _style_header(ws11, 1, len(headers))

        allergen_stats = conn.execute("""
            SELECT da.allergen_slug, COUNT(da.dish_id)
            FROM dish_allergens da
            GROUP BY da.allergen_slug
            ORDER BY COUNT(da.dish_id) DESC
        """).fetchall()

        for i, row_data in enumerate(allergen_stats, 2):
            for col, val in enumerate(row_data, 1):
                ws11.cell(row=i, column=col, value=val)

        _auto_width(ws11)
        _add_borders(ws11, 1, len(allergen_stats) + 1, len(headers))
    except Exception as e:
        ws11.cell(row=2, column=1, value=f"Нет данных: {e}")

    # Save
    output_path = os.path.join(os.path.dirname(__file__), '..', 'data', 'processed', 'pipeline_report.xlsx')
    output_path = os.path.abspath(output_path)
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    wb.save(output_path)

    conn.close()

    print(f"\n  Файл сохранён: {output_path}")
    print(f"  Листов: {len(wb.sheetnames)}")
    print(f"  ({', '.join(wb.sheetnames)})")

    return output_path


if __name__ == '__main__':
    generate_excel_report()
